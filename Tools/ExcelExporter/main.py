import json
import os
from tkinter import filedialog
from typing import Any
from traceback import print_exc

import openpyxl


class DataType:
    name: str

    __datatype_registry: dict[str, type[DataType]] = {}

    def __init__(self, name: str):
        self.name = name

    def parse_value(self, value: str) -> Any:
        """
        将字符串值解析为对应的数据类型
        """
        return value
        
    @staticmethod
    def register(name: str):
        def wrapper(datatype_cls: type[DataType]):
            DataType.__datatype_registry[name] = datatype_cls
            return datatype_cls
        return wrapper
    
    @staticmethod
    def from_definition(definition: str):
        """
        从数据类型定义字符串获取 DataType 实例
        """
        lt_index = definition.find("<")
        if lt_index != -1:
            if not definition.endswith(">"):
                raise ValueError(f"Invalid data type definition: {definition}")
            type_name = definition[:lt_index]
            type_params = [param.strip() for param in definition[lt_index + 1:-1].split(",")]
        else:
            type_name = definition
            type_params = []
        datatype = DataType.__datatype_registry.get(type_name)
        if datatype is None:
            raise ValueError(f"Unsupported data type: {type_name}")
        return datatype(type_name, *type_params)


@DataType.register("int")
class IntType(DataType):
    def __init__(self, name: str):
        super().__init__(name)
    
    def parse_value(self, value: str):
        return int(value)
    

@DataType.register("float")
class FloatType(DataType):
    def __init__(self, name: str):
        super().__init__(name)
    
    def parse_value(self, value: str):
        return float(value)


@DataType.register("string")
class StringType(DataType):
    def __init__(self, name: str):
        super().__init__(name)
    
    def parse_value(self, value: str):
        return value


@DataType.register("id")
class IDType(DataType):
    def __init__(self, name: str):
        super().__init__(name)
    
    def parse_value(self, value: str):
        return value  # ID 类型暂时不进行解析，直接返回字符串值


@DataType.register("path")
class PathType(DataType):
    def __init__(self, name: str):
        super().__init__(name)
    
    def parse_value(self, value: str):
        return value  # 路径类型暂时不进行解析，直接返回字符串值


@DataType.register("ref")
class RefType(DataType):
    ref_sheet: str

    def __init__(self, name: str, ref_sheet: str):
        super().__init__(name)
        self.ref_sheet = ref_sheet
    
    def parse_value(self, value: str):
        return value  # 引用类型暂时不进行解析，直接返回字符串值


class ColumnHeader:
    name: str
    datatype: DataType
    default_value: str

    def __init__(self, name: str, datatype: DataType, default_value: str = ""):
        self.name = name
        self.datatype = datatype
        self.default_value = default_value


class Sheet:
    def __init__(self) -> None:
        self._headers: list[ColumnHeader] = []
        self._raw_data: list[list[str]] = []
    
    def load_from_sheet(self, workbook: openpyxl.Workbook, sheet_name: str):
        """
        从 Excel Sheet 加载数据
        """
        sheet = workbook[sheet_name]
        sheet_iter = sheet.iter_rows(values_only=True)

        try:
            _ = next(sheet_iter)  # 第一行：字段中文名，不作处理，直接跳过
            row1 = next(sheet_iter)  # 第二行：字段英文名
            row2 = next(sheet_iter)  # 第三行：字段数据类型
            row3 = next(sheet_iter)  # 第四行：字段默认值
        except StopIteration:
            raise ValueError(f"Sheet '{sheet_name}' does not have enough rows for headers")

        for name, dtype, default_value in zip(row1, row2, row3):
            if not isinstance(name, str):
                raise ValueError(f"Invalid column name: {name}")
            if not isinstance(dtype, str):
                raise ValueError(f"Invalid data type definition: {dtype}")
            datatype = DataType.from_definition(dtype)
            self._headers.append(ColumnHeader(name, datatype, str(default_value) if default_value is not None else ""))

        for row in sheet_iter:
            row_data = []
            for cell in row:
                if cell is None:
                    cell = ""
                row_data.append(cell)
            self._raw_data.append(row_data)


    def get_row(self, index: int) -> dict[str, Any]:
        """
        获取 Sheet 中指定索引的行数据，返回字典
        """
        if index < 0 or index >= len(self._raw_data):
            raise IndexError(f"Row index {index} is out of bounds")
        
        row = self._raw_data[index]
        row_dict = {}
        for header, cell in zip(self._headers, row):
            if cell == "":
                cell = header.default_value
            parsed_value = header.datatype.parse_value(str(cell))

            # 处理数组类型字段
            bracket_start = header.name.find("[")
            if bracket_start != -1:
                if not header.name.endswith("]"):
                    raise ValueError(f"Invalid array field name: {header.name}")
                base_name = header.name[:bracket_start]
                index = int(header.name[bracket_start + 1:-1])
                arr = row_dict.setdefault(base_name, [])
                if len(arr) <= index:
                    arr.extend([None] * (index - len(arr)) + [parsed_value])
                elif arr[index] is None:
                    arr[index] = parsed_value
                else:
                    raise ValueError(f"Duplicate value for array field '{header.name}'")
            else:
                row_dict[header.name] = parsed_value

        # 验证数组类型字段是否缺失值
        for header_name, cell in row_dict.items():
            if isinstance(cell, list):
                for i, value in enumerate(cell):
                    if value is None:
                        raise ValueError(f"Array field '{header_name}[{i}]' is missed")

        return row_dict
    
    def iter_rows(self):
        """
        迭代 Sheet 中的所有行数据，返回字典
        """
        for index in range(len(self._raw_data)):
            yield self.get_row(index)

    def save_to_json(self, file_path: str):
        """
        将 Sheet 数据保存为 JSON 文件
        """
        try:
            data = list(self.iter_rows())
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except:
            print(f"Failed to save sheet data to JSON file '{file_path}':")
            print_exc()


class SheetDatabase:
    def __init__(self) -> None:
        self._sheets: dict[str, Sheet] = {}
    
    def load_from_excel(self, excel_dir: str):
        """
        从 Excel 加载数据
        """
        for file in os.listdir(excel_dir):
            if file.endswith(".xlsx") and not file.startswith("~$"):
                self._load_excel_sheets(os.path.join(excel_dir, file))
        
    def _load_excel_sheets(self, file_path: str):
        """
        加载 Excel 文件中的所有 Sheet
        """
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = Sheet()
            try:
                sheet.load_from_sheet(workbook, sheet_name)
                self._sheets[sheet_name] = sheet
            except:
                print(f"Failed to load sheet '{sheet_name}' from file '{file_path}':")
                print_exc()
    
    def save_to_json(self, output_dir: str):
        """
        将所有 Sheet 数据保存为 JSON 文件
        """
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        for sheet_name, sheet in self._sheets.items():
            json_file_path = os.path.join(output_dir, f"{sheet_name}.json")
            sheet.save_to_json(json_file_path)


def main():
    config = {
        "output_dir": "./test_output"
    }

    if os.path.exists("local_config.json"):
        with open("local_config.json", "r", encoding="utf-8") as f:
            local_config = json.load(f)
            config.update(local_config)

    if "excel_dir" not in config or not os.path.exists(config["excel_dir"]):
        excel_dir = filedialog.askdirectory(title="选择数据表文件夹", mustexist=True, initialdir=os.getcwd())
        if not excel_dir:
            return
        config["excel_dir"] = os.path.abspath(excel_dir)
    
    with open("local_config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)
    
    database = SheetDatabase()
    database.load_from_excel(config["excel_dir"])
    database.save_to_json(config["output_dir"])


if __name__ == "__main__":
    main()
